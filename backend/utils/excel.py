from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_
from models.models import Horario, Docente, Asignatura, Modulo, Carrera, Nivel, CargaHoraria, PeriodoAcademico
import io


# ─── ESTILOS ──────────────────────────────────────────────────────────────────

def borde_completo():
    lado = Side(style="thin", color="000000")
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def celda_titulo(ws, fila, col_inicio, col_fin, texto, color="1F3864", size=12):
    ws.merge_cells(start_row=fila, start_column=col_inicio, end_row=fila, end_column=col_fin)
    c = ws.cell(row=fila, column=col_inicio, value=texto)
    c.font = Font(bold=True, color="FFFFFF", size=size)
    c.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return c

def celda_subtitulo(ws, fila, col_inicio, col_fin, texto, color="2E75B6"):
    ws.merge_cells(start_row=fila, start_column=col_inicio, end_row=fila, end_column=col_fin)
    c = ws.cell(row=fila, column=col_inicio, value=texto)
    c.font = Font(bold=True, color="FFFFFF", size=11)
    c.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="center")
    return c

def celda_encabezado(cell, color="BDD7EE"):
    cell.font = Font(bold=True, size=10, color="1F3864")
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = borde_completo()

def celda_dato(cell, fila_par=False):
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = borde_completo()
    cell.font = Font(size=10)
    if fila_par:
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

def celda_centro(cell, fila_par=False):
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = borde_completo()
    cell.font = Font(size=10)
    if fila_par:
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

def autoajustar(ws, anchos: dict):
    for col, ancho in anchos.items():
        ws.column_dimensions[get_column_letter(col)].width = ancho


# ─── CARGAR DATOS RELACIONADOS ────────────────────────────────────────────────

async def cargar_catalogos(db: AsyncSession):
    docentes_map = {}
    r = await db.execute(select(Docente))
    for d in r.scalars().all():
        docentes_map[d.id] = f"{d.nombre} {d.apellido}"

    asignaturas_map = {}
    r = await db.execute(select(Asignatura))
    for a in r.scalars().all():
        asignaturas_map[a.id] = a

    carreras_map = {}
    sedes_map = {}
    r = await db.execute(select(Carrera))
    for c in r.scalars().all():
        carreras_map[c.id] = c.nombre
        sedes_map[c.id] = c.sede or 'Quito'

    niveles_map = {}
    r = await db.execute(select(Nivel))
    for n in r.scalars().all():
        niveles_map[n.id] = n.numero

    return docentes_map, asignaturas_map, carreras_map, niveles_map, sedes_map


# ─── HOJA MÓDULO (formato ITQ) ────────────────────────────────────────────────

async def escribir_hoja_modulo(ws, modulo: Modulo, periodo: PeriodoAcademico, horarios: list, db: AsyncSession):
    docentes_map, asignaturas_map, carreras_map, niveles_map, sedes_map = await cargar_catalogos(db)

    # Ajuste de alturas
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 18

    # Encabezado institucional
    celda_titulo(ws, 1, 1, 10, f"CARRERA: {carreras_map.get(horarios[0].carrera_id, 'ITQ') if horarios else 'ITQ'}", "1F3864", 13)
    celda_titulo(ws, 2, 1, 10, "PLANIFICACIÓN ACADÉMICA", "1F3864", 11)
    celda_titulo(ws, 3, 1, 10, f"PERIODO: {periodo.nombre}", "1F3864", 11)
    celda_titulo(ws, 4, 1, 10, f"{modulo.nombre.upper()}  ({modulo.fecha_inicio} al {modulo.fecha_fin})", "2E75B6", 12)

    fila_actual = 6

    # Separar por jornada
    for jornada_label, jornada_key, horario_txt in [
        ("MODALIDAD PRESENCIAL QUITO — MATUTINO (LUNES A VIERNES)", "matutina", "08H00 - 12H00"),
        ("MODALIDAD PRESENCIAL QUITO — NOCTURNO (LUNES A VIERNES)", "nocturna", "18H30 - 21H30"),
    ]:
        horarios_jornada = [h for h in horarios if h.jornada == jornada_key]
        if not horarios_jornada:
            continue

        # Subtítulo jornada
        ws.row_dimensions[fila_actual].height = 18
        celda_subtitulo(ws, fila_actual, 1, 10, jornada_label, "2E75B6")
        fila_actual += 1

        # Encabezados columnas
        ws.row_dimensions[fila_actual].height = 30
        encabezados = ["MÓDULO", "FECHA INICIO", "FECHA FIN", "ASIGNATURA", "NIVEL", "PARALELO", "HORAS", "HORARIO", "DOCENTE", "OBSERVACIÓN"]
        for col, enc in enumerate(encabezados, 1):
            celda_encabezado(ws.cell(row=fila_actual, column=col, value=enc))
        fila_actual += 1

        # Datos
        for idx, h in enumerate(horarios_jornada):
            ws.row_dimensions[fila_actual].height = 18
            asig = asignaturas_map.get(h.asignatura_id)
            asig_nombre = asig.nombre if asig else "N/A"
            nivel_num = niveles_map.get(h.nivel_id, "")
            docente_nombre = docentes_map.get(h.docente_id, "POR DEFINIR")
            horas = asig.horas_modulo if asig else 0
            par = idx % 2 == 0

            datos = [
                modulo.numero,
                str(modulo.fecha_inicio),
                str(modulo.fecha_fin),
                asig_nombre,
                nivel_num,
                h.paralelo,
                horas,
                f"{h.hora_inicio} - {h.hora_fin}",
                docente_nombre.upper(),
                h.observaciones or "",
            ]

            for col, val in enumerate(datos, 1):
                c = ws.cell(row=fila_actual, column=col, value=val)
                if col in [1, 5, 6, 7]:
                    celda_centro(c, par)
                else:
                    celda_dato(c, par)

            fila_actual += 1

        fila_actual += 1  # Espacio entre jornadas

    # Anchos de columna
    autoajustar(ws, {1: 10, 2: 14, 3: 14, 4: 35, 5: 8, 6: 10, 7: 8, 8: 16, 9: 25, 10: 20})


# ─── HOJA CARGA HORARIA POR MÓDULO ────────────────────────────────────────────

async def escribir_hoja_carga_modulos(ws, periodo: PeriodoAcademico, modulos: list, db: AsyncSession):
    docentes_map, _, carreras_map, _, sedes_map = await cargar_catalogos(db)

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 18

    # Obtener nombre carrera del período
    r = await db.execute(select(Horario).where(Horario.periodo_id == periodo.id).limit(1))
    primer_horario = r.scalar_one_or_none()
    carrera_nombre = carreras_map.get(primer_horario.carrera_id, "ITQ") if primer_horario else "ITQ"

    celda_titulo(ws, 1, 1, 8, f"CARRERA: {carrera_nombre}", "1F3864", 13)
    celda_titulo(ws, 2, 1, 8, "PLANIFICACIÓN ACADÉMICA DOCENTES A TIEMPO COMPLETO & PARCIAL", "1F3864", 11)
    celda_titulo(ws, 3, 1, 8, f"PERIODO: {periodo.nombre}", "1F3864", 11)

    fila = 5
    # Encabezados
    encabezados = ["DOCENTES", "TIPO", "MÓDULO 1\n(asig)", "MÓDULO 2\n(asig)", "MÓDULO 3\n(asig)", "TOTAL\nASIG", "HORAS TC", "HORAS TP"]
    ws.row_dimensions[fila].height = 35
    for col, enc in enumerate(encabezados, 1):
        celda_encabezado(ws.cell(row=fila, column=col, value=enc))
    fila += 1

    r = await db.execute(select(Docente).where(Docente.activo == True).order_by(Docente.apellido))
    docentes = r.scalars().all()

    for idx, docente in enumerate(docentes):
        ws.row_dimensions[fila].height = 18
        par = idx % 2 == 0
        nombre = f"{docente.nombre} {docente.apellido}".upper()
        tipo = "TC" if docente.tipo.value == "tiempo_completo" else "TP"

        c1 = ws.cell(row=fila, column=1, value=nombre)
        celda_dato(c1, par)
        c2 = ws.cell(row=fila, column=2, value=tipo)
        celda_centro(c2, par)

        total_asig = 0
        total_horas_tc = 0
        total_horas_tp = 0

        for i, modulo in enumerate(modulos[:3], 3):
            r2 = await db.execute(
                select(CargaHoraria).where(and_(
                    CargaHoraria.docente_id == docente.id,
                    CargaHoraria.modulo_id == modulo.id,
                ))
            )
            carga = r2.scalar_one_or_none()
            asig = carga.total_asignaturas if carga else 0
            horas = carga.total_horas if carga else 0
            total_asig += asig
            if tipo == "TC":
                total_horas_tc += horas
            else:
                total_horas_tp += horas

            c = ws.cell(row=fila, column=i, value=asig)
            celda_centro(c, par)
            if asig >= 3:
                c.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

        ct = ws.cell(row=fila, column=6, value=total_asig)
        celda_centro(ct, par)
        ct.font = Font(bold=True, size=10)

        ctc = ws.cell(row=fila, column=7, value=round(total_horas_tc, 1))
        celda_centro(ctc, par)

        ctp = ws.cell(row=fila, column=8, value=round(total_horas_tp, 1))
        celda_centro(ctp, par)

        fila += 1

    autoajustar(ws, {1: 30, 2: 8, 3: 14, 4: 14, 5: 14, 6: 10, 7: 10, 8: 10})


# ─── HOJA CARGA TOTAL ─────────────────────────────────────────────────────────

async def escribir_hoja_carga_total(ws, periodo: PeriodoAcademico, db: AsyncSession):
    docentes_map, _, carreras_map, _, sedes_map = await cargar_catalogos(db)

    ws.row_dimensions[1].height = 22

    r = await db.execute(select(Horario).where(Horario.periodo_id == periodo.id).limit(1))
    primer_horario = r.scalar_one_or_none()
    carrera_nombre = carreras_map.get(primer_horario.carrera_id, "ITQ") if primer_horario else "ITQ"

    celda_titulo(ws, 1, 1, 7, f"CARRERA: {carrera_nombre}", "1F3864", 13)
    celda_titulo(ws, 2, 1, 7, "DISTRIBUCIÓN DE CARGA HORARIA DE DOCENCIA TIEMPO COMPLETO", "1F3864", 11)
    celda_titulo(ws, 3, 1, 7, f"PERIODO: {periodo.nombre}", "1F3864", 11)

    fila = 5
    encabezados = ["DOCENTES", "MÓDULO 1", "MÓDULO 2", "MÓDULO 3", "TOTAL", "MÍNIMO", "ESTADO"]
    ws.row_dimensions[fila].height = 30
    for col, enc in enumerate(encabezados, 1):
        celda_encabezado(ws.cell(row=fila, column=col, value=enc))

    # Fila de referencia mínimo/máximo
    fila += 1
    ws.cell(row=fila, column=6, value=272).font = Font(bold=True, color="FF0000", size=10)
    ws.cell(row=fila, column=6).alignment = Alignment(horizontal="center")
    fila += 1

    r = await db.execute(
        select(Docente)
        .where(Docente.activo == True)
        .order_by(Docente.apellido)
    )
    docentes = r.scalars().all()

    r2 = await db.execute(
        select(Modulo)
        .where(Modulo.periodo_id == periodo.id)
        .order_by(Modulo.numero)
    )
    modulos = r2.scalars().all()

    for idx, docente in enumerate(docentes):
        ws.row_dimensions[fila].height = 18
        par = idx % 2 == 0
        nombre = f"{docente.nombre} {docente.apellido}".upper()
        es_tc = docente.tipo.value == "tiempo_completo"

        c1 = ws.cell(row=fila, column=1, value=nombre)
        celda_dato(c1, par)

        total_horas = 0
        for i, modulo in enumerate(modulos[:3], 2):
            r3 = await db.execute(
                select(CargaHoraria).where(and_(
                    CargaHoraria.docente_id == docente.id,
                    CargaHoraria.modulo_id == modulo.id,
                ))
            )
            carga = r3.scalar_one_or_none()
            horas = round(carga.total_horas, 1) if carga else 0
            total_horas += horas
            c = ws.cell(row=fila, column=i, value=horas)
            celda_centro(c, par)

        ct = ws.cell(row=fila, column=5, value=round(total_horas, 1))
        celda_centro(ct, par)
        ct.font = Font(bold=True, size=10)

        ws.cell(row=fila, column=6, value=272 if es_tc else "N/A").border = borde_completo()

        if es_tc:
            if total_horas < 272:
                estado = "⚠ BAJO MÍNIMO"
                color_txt = "FF0000"
            elif total_horas > 380:
                estado = "⚠ EXCEDE MÁXIMO"
                color_txt = "FF0000"
            else:
                estado = "✓ OK"
                color_txt = "375623"
        else:
            estado = "TIEMPO PARCIAL"
            color_txt = "2E75B6"

        ce = ws.cell(row=fila, column=7, value=estado)
        ce.font = Font(bold=True, color=color_txt, size=10)
        ce.border = borde_completo()
        ce.alignment = Alignment(horizontal="center", vertical="center")

        fila += 1

    autoajustar(ws, {1: 30, 2: 12, 3: 12, 4: 12, 5: 10, 6: 10, 7: 18})


# ─── FUNCIÓN PRINCIPAL ────────────────────────────────────────────────────────

async def generar_excel(periodo_id: str, db: AsyncSession) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    r = await db.execute(select(PeriodoAcademico).where(PeriodoAcademico.id == periodo_id))
    periodo = r.scalar_one_or_none()
    if not periodo:
        raise Exception("Período no encontrado")

    r = await db.execute(
        select(Modulo)
        .where(Modulo.periodo_id == periodo_id)
        .order_by(Modulo.numero)
    )
    modulos = r.scalars().all()
    if not modulos:
        raise Exception("No hay módulos registrados para este período")

    # Obtener todas las carreras activas del periodo
    r = await db.execute(
        select(Carrera).where(Carrera.activo == True).order_by(Carrera.nombre, Carrera.sede)
    )
    todas_carreras = r.scalars().all()

    # Agrupar carreras por nombre (cada nombre = una carrera con posibles varias sedes)
    from collections import OrderedDict
    carreras_agrupadas = OrderedDict()
    for c in todas_carreras:
        nombre_lower = c.nombre.lower()
        if nombre_lower not in carreras_agrupadas:
            carreras_agrupadas[nombre_lower] = []
        carreras_agrupadas[nombre_lower].append(c)

    docentes_map, asignaturas_map, carreras_map, niveles_map, sedes_map = await cargar_catalogos(db)

    # Hojas por módulo — una hoja por carrera con secciones por sede
    for modulo in modulos[:3]:
        for nombre_carrera, sedes in carreras_agrupadas.items():
            # Verificar si hay horarios para esta carrera en este modulo
            ids_sedes = [c.id for c in sedes]
            r = await db.execute(
                select(Horario)
                .where(
                    Horario.modulo_id == modulo.id,
                    Horario.carrera_id.in_(ids_sedes)
                )
                .order_by(Horario.jornada, Horario.nivel_id, Horario.paralelo, Horario.hora_inicio)
            )
            horarios_carrera = r.scalars().all()
            if not horarios_carrera:
                continue

            # Titulo de hoja: MODULO N - Carrera (max 31 chars para Excel)
            titulo_hoja = f"M{modulo.numero}-{sedes[0].nombre[:20]}"
            ws = wb.create_sheet(title=titulo_hoja)

            # Encabezado general
            celda_titulo(ws, 1, 1, 10, f"CARRERA: {sedes[0].nombre.upper()}", "1F3864", 13)
            celda_titulo(ws, 2, 1, 10, "PLANIFICACION ACADEMICA", "1F3864", 11)
            celda_titulo(ws, 3, 1, 10, f"PERIODO: {periodo.nombre}", "1F3864", 11)
            celda_titulo(ws, 4, 1, 10,
                f"{modulo.nombre.upper()}  ({modulo.fecha_inicio} al {modulo.fecha_fin})",
                "2E75B6", 12)

            fila = 6

            for sede_carrera in sorted(sedes, key=lambda c: c.sede or "Quito"):
                horarios_sede = [h for h in horarios_carrera if h.carrera_id == sede_carrera.id]
                if not horarios_sede:
                    continue

                # Titulo sede
                ws.row_dimensions[fila].height = 20
                celda_subtitulo(ws, fila, 1, 10,
                    f"SEDE: {(sede_carrera.sede or 'Quito').upper()}", "375623")
                fila += 1

                for jornada_label, jornada_key, horario_txt in [
                    ("MODALIDAD PRESENCIAL — MATUTINO (LUNES A VIERNES)", "matutina", "08H00 - 12H00"),
                    ("MODALIDAD PRESENCIAL — NOCTURNO (LUNES A VIERNES)", "nocturna", "18H30 - 21H30"),
                ]:
                    horarios_jornada = [h for h in horarios_sede if h.jornada == jornada_key]
                    if not horarios_jornada:
                        continue

                    ws.row_dimensions[fila].height = 18
                    celda_subtitulo(ws, fila, 1, 10,
                        f"{jornada_label} — {horario_txt}", "2E75B6")
                    fila += 1

                    ws.row_dimensions[fila].height = 30
                    encabezados = ["MODULO","FECHA INICIO","FECHA FIN","ASIGNATURA","NIVEL","PARALELO","HORAS","HORARIO","DOCENTE","OBSERVACION"]
                    for col, enc in enumerate(encabezados, 1):
                        celda_encabezado(ws.cell(row=fila, column=col, value=enc))
                    fila += 1

                    for idx, h in enumerate(horarios_jornada):
                        ws.row_dimensions[fila].height = 18
                        asig = asignaturas_map.get(h.asignatura_id)
                        par = idx % 2 == 0
                        datos = [
                            modulo.numero,
                            str(modulo.fecha_inicio),
                            str(modulo.fecha_fin),
                            asig.nombre if asig else "N/A",
                            niveles_map.get(h.nivel_id, ""),
                            h.paralelo,
                            asig.horas_modulo if asig else 0,
                            f"{h.hora_inicio} - {h.hora_fin}",
                            docentes_map.get(h.docente_id, "POR DEFINIR").upper(),
                            h.observaciones or "",
                        ]
                        for col, val in enumerate(datos, 1):
                            c = ws.cell(row=fila, column=col, value=val)
                            if col in [1, 5, 6, 7]:
                                celda_centro(c, par)
                            else:
                                celda_dato(c, par)
                        fila += 1
                    fila += 1

                fila += 1

            autoajustar(ws, {1:10, 2:14, 3:14, 4:35, 5:8, 6:10, 7:8, 8:16, 9:25, 10:20})

    # Hoja carga por módulos
    ws_carga = wb.create_sheet(title="Carga horaria por modulos")
    await escribir_hoja_carga_modulos(ws_carga, periodo, modulos, db)

    # Hoja carga total
    ws_total = wb.create_sheet(title="Carga horaria total")
    await escribir_hoja_carga_total(ws_total, periodo, db)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()

# ─── REPORTE POR DOCENTE ──────────────────────────────────────────────────────

async def generar_excel_por_docente(periodo_id: str, docente_id: str, db: AsyncSession) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    r = await db.execute(select(PeriodoAcademico).where(PeriodoAcademico.id == periodo_id))
    periodo = r.scalar_one_or_none()
    if not periodo:
        raise Exception("Período no encontrado")

    r = await db.execute(select(Docente).where(Docente.id == docente_id))
    docente = r.scalar_one_or_none()
    if not docente:
        raise Exception("Docente no encontrado")

    r = await db.execute(
        select(Modulo).where(Modulo.periodo_id == periodo_id).order_by(Modulo.numero)
    )
    modulos = r.scalars().all()

    docentes_map, asignaturas_map, carreras_map, niveles_map, sedes_map = await cargar_catalogos(db)

    # Hoja por cada módulo
    for modulo in modulos:
        ws = wb.create_sheet(title=f"MÓDULO {modulo.numero}")
        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 18
        ws.row_dimensions[3].height = 18
        ws.row_dimensions[4].height = 18

        celda_titulo(ws, 1, 1, 9, f"DOCENTE: {docente.nombre} {docente.apellido}".upper(), "1F3864", 13)
        celda_titulo(ws, 2, 1, 9, f"TIPO: {docente.tipo.value.replace('_', ' ').upper()}", "1F3864", 11)
        celda_titulo(ws, 3, 1, 9, f"PERIODO: {periodo.nombre}", "1F3864", 11)
        celda_titulo(ws, 4, 1, 9, f"{modulo.nombre.upper()} ({modulo.fecha_inicio} al {modulo.fecha_fin})", "2E75B6", 12)

        r = await db.execute(
            select(Horario)
            .where(Horario.modulo_id == modulo.id, Horario.docente_id == docente_id)
            .order_by(Horario.jornada, Horario.dia, Horario.hora_inicio)
        )
        horarios = r.scalars().all()

        fila = 6
        for jornada_key, jornada_label, horario_txt in [
            ("matutina", "MATUTINO (LUNES A VIERNES)", "08:00 - 12:00"),
            ("nocturna", "NOCTURNO (LUNES A VIERNES)", "18:30 - 21:30"),
        ]:
            h_jornada = [h for h in horarios if h.jornada == jornada_key]
            if not h_jornada:
                continue

            ws.row_dimensions[fila].height = 18
            celda_subtitulo(ws, fila, 1, 9, f"MODALIDAD PRESENCIAL — {jornada_label}", "2E75B6")
            fila += 1

            encabezados = ["MÓDULO", "FECHA INICIO", "FECHA FIN", "ASIGNATURA", "NIVEL", "PARALELO", "HORARIO", "CARRERA", "OBSERVACIÓN"]
            ws.row_dimensions[fila].height = 30
            for col, enc in enumerate(encabezados, 1):
                celda_encabezado(ws.cell(row=fila, column=col, value=enc))
            fila += 1

            for idx, h in enumerate(h_jornada):
                par = idx % 2 == 0
                asig = asignaturas_map.get(h.asignatura_id)
                datos = [
                    modulo.numero,
                    modulo.fecha_inicio,
                    modulo.fecha_fin,
                    asig.nombre if asig else "N/A",
                    niveles_map.get(h.nivel_id, ""),
                    h.paralelo,
                    f"{h.hora_inicio} - {h.hora_fin}",
                    f"{carreras_map.get(h.carrera_id, 'N/A')} — {sedes_map.get(h.carrera_id, 'Quito')}",
                    h.observaciones or "",
                ]
                for col, val in enumerate(datos, 1):
                    c = ws.cell(row=fila, column=col, value=val)
                    if col in [1, 2, 3, 5, 6]:
                        celda_centro(c, par)
                    else:
                        celda_dato(c, par)
                fila += 1
            fila += 1

        autoajustar(ws, {1: 10, 2: 14, 3: 14, 4: 35, 5: 8, 6: 10, 7: 16, 8: 30, 9: 20})

    # Hojas de carga horaria usando las funciones estandar
    ws_carga = wb.create_sheet(title="Carga horaria por modulos")
    await escribir_hoja_carga_modulos(ws_carga, periodo, modulos, db)

    ws_total = wb.create_sheet(title="Carga horaria total")
    await escribir_hoja_carga_total(ws_total, periodo, db)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


# ─── REPORTE POR CARRERA ──────────────────────────────────────────────────────

async def generar_excel_por_carrera(periodo_id: str, carrera_id: str, db: AsyncSession) -> bytes:
    """
    Genera reporte por carrera incluyendo todas las sedes que tengan el mismo nombre.
    Cada sede aparece como seccion separada dentro de cada hoja de modulo.
    """
    wb = Workbook()
    wb.remove(wb.active)

    r = await db.execute(select(PeriodoAcademico).where(PeriodoAcademico.id == periodo_id))
    periodo = r.scalar_one_or_none()
    if not periodo:
        raise Exception("Periodo no encontrado")

    r = await db.execute(select(Carrera).where(Carrera.id == carrera_id))
    carrera = r.scalar_one_or_none()
    if not carrera:
        raise Exception("Carrera no encontrada")

    # Buscar todas las sedes de esta misma carrera (mismo nombre, case-insensitive)
    from sqlalchemy import func as sqlfunc
    r = await db.execute(
        select(Carrera).where(
            sqlfunc.lower(Carrera.nombre) == carrera.nombre.lower(),
            Carrera.activo == True
        )
    )
    todas_sedes = r.scalars().all()
    ids_sedes = [c.id for c in todas_sedes]

    r = await db.execute(
        select(Modulo).where(Modulo.periodo_id == periodo_id).order_by(Modulo.numero)
    )
    modulos = r.scalars().all()

    docentes_map, asignaturas_map, carreras_map, niveles_map, sedes_map = await cargar_catalogos(db)

    # Hojas por modulo - incluye todas las sedes separadas por seccion
    for modulo in modulos:
        ws = wb.create_sheet(title=f"MODULO {modulo.numero}")

        # Titulo general
        celda_titulo(ws, 1, 1, 10, f"CARRERA: {carrera.nombre.upper()}", "1F3864", 13)
        celda_titulo(ws, 2, 1, 10, "PLANIFICACION ACADEMICA — TODAS LAS SEDES", "1F3864", 11)
        celda_titulo(ws, 3, 1, 10, f"PERIODO: {periodo.nombre}", "1F3864", 11)
        celda_titulo(ws, 4, 1, 10,
            f"{modulo.nombre.upper()}  ({modulo.fecha_inicio} al {modulo.fecha_fin})",
            "2E75B6", 12)

        fila = 6

        for sede_carrera in sorted(todas_sedes, key=lambda c: c.sede or 'Quito'):
            # Obtener horarios de esta sede
            r = await db.execute(
                select(Horario)
                .where(Horario.modulo_id == modulo.id, Horario.carrera_id == sede_carrera.id)
                .order_by(Horario.jornada, Horario.nivel_id, Horario.paralelo, Horario.hora_inicio)
            )
            horarios_sede = r.scalars().all()
            if not horarios_sede:
                continue

            # Titulo de sede
            ws.row_dimensions[fila].height = 20
            celda_subtitulo(ws, fila, 1, 10,
                f"SEDE: {(sede_carrera.sede or 'Quito').upper()}", "375623")
            fila += 1

            for jornada_label, jornada_key, horario_txt in [
                ("MODALIDAD PRESENCIAL — MATUTINO (LUNES A VIERNES)", "matutina", "08H00 - 12H00"),
                ("MODALIDAD PRESENCIAL — NOCTURNO (LUNES A VIERNES)", "nocturna", "18H30 - 21H30"),
            ]:
                horarios_jornada = [h for h in horarios_sede if h.jornada == jornada_key]
                if not horarios_jornada:
                    continue

                ws.row_dimensions[fila].height = 18
                celda_subtitulo(ws, fila, 1, 10, f"{jornada_label} — {horario_txt}", "2E75B6")
                fila += 1

                ws.row_dimensions[fila].height = 30
                encabezados = ["MODULO","FECHA INICIO","FECHA FIN","ASIGNATURA","NIVEL","PARALELO","HORAS","HORARIO","DOCENTE","OBSERVACION"]
                for col, enc in enumerate(encabezados, 1):
                    celda_encabezado(ws.cell(row=fila, column=col, value=enc))
                fila += 1

                for idx, h in enumerate(horarios_jornada):
                    ws.row_dimensions[fila].height = 18
                    asig = asignaturas_map.get(h.asignatura_id)
                    par = idx % 2 == 0
                    datos = [
                        modulo.numero,
                        str(modulo.fecha_inicio),
                        str(modulo.fecha_fin),
                        asig.nombre if asig else "N/A",
                        niveles_map.get(h.nivel_id, ""),
                        h.paralelo,
                        asig.horas_modulo if asig else 0,
                        f"{h.hora_inicio} - {h.hora_fin}",
                        docentes_map.get(h.docente_id, "POR DEFINIR").upper(),
                        h.observaciones or "",
                    ]
                    for col, val in enumerate(datos, 1):
                        c = ws.cell(row=fila, column=col, value=val)
                        if col in [1, 5, 6, 7]:
                            celda_centro(c, par)
                        else:
                            celda_dato(c, par)
                    fila += 1

                fila += 1  # Espacio entre jornadas

            fila += 1  # Espacio entre sedes

        autoajustar(ws, {1:10, 2:14, 3:14, 4:35, 5:8, 6:10, 7:8, 8:16, 9:25, 10:20})

    # Carga horaria - docentes que tienen clases en cualquiera de las sedes
    r = await db.execute(
        select(Docente.id).join(Horario, Horario.docente_id == Docente.id)
        .where(Horario.periodo_id == periodo_id, Horario.carrera_id.in_(ids_sedes))
        .distinct()
    )
    docente_ids_carrera = [row[0] for row in r.fetchall()]

    ws_carga = wb.create_sheet(title="Carga horaria por modulos")
    await escribir_hoja_carga_modulos(ws_carga, periodo, modulos, db)

    ws_total = wb.create_sheet(title="Carga horaria total")
    await escribir_hoja_carga_total(ws_total, periodo, db)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


# ─── REPORTE POR NIVEL ────────────────────────────────────────────────────────

async def generar_excel_por_nivel(periodo_id: str, nivel_id: str, db: AsyncSession) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    r = await db.execute(select(PeriodoAcademico).where(PeriodoAcademico.id == periodo_id))
    periodo = r.scalar_one_or_none()
    if not periodo:
        raise Exception("Período no encontrado")

    r = await db.execute(select(Nivel).where(Nivel.id == nivel_id))
    nivel = r.scalar_one_or_none()
    if not nivel:
        raise Exception("Nivel no encontrado")

    # Buscar niveles equivalentes en todas las sedes (mismo numero, misma carrera por nombre)
    r = await db.execute(select(Carrera).where(Carrera.id == nivel.carrera_id))
    carrera_nivel = r.scalar_one_or_none()

    from sqlalchemy import func as sqlfunc
    r = await db.execute(
        select(Carrera).where(
            sqlfunc.lower(Carrera.nombre) == carrera_nivel.nombre.lower(),
            Carrera.activo == True
        )
    )
    todas_sedes_carrera = r.scalars().all()
    ids_sedes_carrera = [c.id for c in todas_sedes_carrera]

    # Niveles del mismo numero en todas las sedes
    r = await db.execute(
        select(Nivel).where(
            Nivel.numero == nivel.numero,
            Nivel.carrera_id.in_(ids_sedes_carrera)
        )
    )
    niveles_todas_sedes = r.scalars().all()
    ids_niveles = [n.id for n in niveles_todas_sedes]

    r = await db.execute(
        select(Modulo).where(Modulo.periodo_id == periodo_id).order_by(Modulo.numero)
    )
    modulos = r.scalars().all()
    docentes_map, asignaturas_map, carreras_map, niveles_map, sedes_map = await cargar_catalogos(db)

    # Hojas por módulo
    for modulo in modulos:
        ws = wb.create_sheet(title=f"MÓDULO {modulo.numero}")
        ws.row_dimensions[1].height = 22
        ws.row_dimensions[4].height = 18

        celda_titulo(ws, 1, 1, 9, f"NIVEL: {nivel.nombre or f'Nivel {nivel.numero}'}", "1F3864", 13)
        celda_titulo(ws, 2, 1, 9, f"PERIODO: {periodo.nombre}", "1F3864", 11)
        celda_titulo(ws, 3, 1, 9, f"{modulo.nombre.upper()} ({modulo.fecha_inicio} al {modulo.fecha_fin})", "2E75B6", 12)

        r = await db.execute(
            select(Horario)
            .where(Horario.modulo_id == modulo.id, Horario.nivel_id.in_(ids_niveles))
            .order_by(Horario.carrera_id, Horario.jornada, Horario.paralelo, Horario.hora_inicio)
        )
        horarios = r.scalars().all()

        # Agrupar por sede
        sedes_en_nivel = {}
        for h in horarios:
            sede = sedes_map.get(h.carrera_id, 'Quito')
            if sede not in sedes_en_nivel:
                sedes_en_nivel[sede] = []
            sedes_en_nivel[sede].append(h)

        fila = 5
        for sede_nombre in sorted(sedes_en_nivel.keys()):
            horarios_sede = sedes_en_nivel[sede_nombre]

            # Titulo sede
            ws.row_dimensions[fila].height = 20
            celda_subtitulo(ws, fila, 1, 9, f"SEDE: {sede_nombre.upper()}", "375623")
            fila += 1

            for jornada_key, jornada_label in [
                ("matutina", "MATUTINO (LUNES A VIERNES)"),
                ("nocturna", "NOCTURNO (LUNES A VIERNES)"),
            ]:
                h_jornada = [h for h in horarios_sede if h.jornada == jornada_key]
                if not h_jornada:
                    continue

                celda_subtitulo(ws, fila, 1, 9, f"MODALIDAD PRESENCIAL — {jornada_label}", "2E75B6")
                fila += 1

                encabezados = ["MODULO", "FECHA INICIO", "FECHA FIN", "ASIGNATURA", "PARALELO", "HORAS", "HORARIO", "DOCENTE", "OBSERVACION"]
                ws.row_dimensions[fila].height = 30
                for col, enc in enumerate(encabezados, 1):
                    celda_encabezado(ws.cell(row=fila, column=col, value=enc))
                fila += 1

                for idx, h in enumerate(h_jornada):
                    par = idx % 2 == 0
                    asig = asignaturas_map.get(h.asignatura_id)
                    datos = [
                        modulo.numero,
                        modulo.fecha_inicio,
                        modulo.fecha_fin,
                        asig.nombre if asig else "N/A",
                        h.paralelo,
                        asig.horas_modulo if asig else 0,
                        f"{h.hora_inicio} - {h.hora_fin}",
                        docentes_map.get(h.docente_id, "POR DEFINIR").upper(),
                        h.observaciones or "",
                    ]
                    for col, val in enumerate(datos, 1):
                        c = ws.cell(row=fila, column=col, value=val)
                        if col in [1, 2, 3, 5, 6]:
                            celda_centro(c, par)
                        else:
                            celda_dato(c, par)
                        if col in [2, 3]:
                            c.number_format = 'DD/MM/YYYY'
                    fila += 1
                fila += 1  # espacio entre jornadas

            fila += 1  # espacio entre sedes

        autoajustar(ws, {1: 10, 2: 14, 3: 14, 4: 35, 5: 10, 6: 8, 7: 16, 8: 25, 9: 20})

    # Hojas de carga horaria usando las funciones estandar
    ws_carga = wb.create_sheet(title="Carga horaria por modulos")
    await escribir_hoja_carga_modulos(ws_carga, periodo, modulos, db)

    ws_total = wb.create_sheet(title="Carga horaria total")
    await escribir_hoja_carga_total(ws_total, periodo, db)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()